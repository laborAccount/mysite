#-*- coding: utf-8 -*-
import os
from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from main.settings import UPLOAD_DIR, TEMP_UPLOAD_DIR
from django.conf import settings
from django.views.decorators.csrf import csrf_protect, csrf_exempt

#openpyxl
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font,Alignment,Border,Side,PatternFill

import urllib.parse
import math

def file_upload(request):

    print(request, request.POST)
    filelist = list()
    fileObject = request.FILES
    temp_up_dir = os.path.normcase(UPLOAD_DIR)

    for i in range(0, len(fileObject)):
        file_dict = dict()
        item = fileObject.get('file[' + str(i) + ']')
        file_dict['item'] = str(item).strip()
        fileOriginName = str(item).strip()
        fileExtension = fileOriginName.split(".")[-1]
        filelist.append(file_dict)

        if not os.path.isdir(temp_up_dir):
            os.makedirs(temp_up_dir)

        fp = open(temp_up_dir + str(fileOriginName), 'wb')
        for chunk in item.chunks():
            fp.write(chunk)
            fp.close()

    return JsonResponse({'return': 'sucess'})

def file_delete(request):
    filename = request.POST.getlist('filename[]')
    temp_up_dir = os.path.normcase(UPLOAD_DIR)

    for name in filename:
        if len(os.listdir(temp_up_dir)) == 0:
            pass
        else:
            for datas in os.listdir(temp_up_dir):
                os.remove(temp_up_dir+datas)

    return JsonResponse({'return': 'sucess'})

def excel_create(title,columns,datalist):
    os.putenv('NLS_LANG','.UTF8')
    file_name = title['file_name']
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename='+urllib.parse.quote(file_name.encode('utf-8'))+'.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title['sheet_name']
    row_num = 0
    ws.merge_cells('B2:E2')
    ws['B2'] = title['title']
    c = ws['B2']
    c.alignment = Alignment(horizontal='center',vertical='center')
    columns = columns
    double_border = Border(left=Side(style='double'),
                     right=Side(style='double'),
                     top=Side(style='double'),
                     bottom=Side(style='double'))

    for col_num in range(0,len(columns)):
        c = ws.cell(row=row_num + 4, column=col_num + 2)
        c.value = columns[col_num][0]
        c.font = Font(name='맑은 고딕',size=11,bold=True,color='FFFFFF')
        c.fill = PatternFill(fgColor="C0C0C0", fill_type = "solid")
        c.alignment = Alignment(horizontal ='center',vertical ='center')
        c.border = double_border
        # set column width
        ws.column_dimensions[get_column_letter(col_num+2)].width = columns[col_num][1]

    for data in datalist:
        row_num += 1
        row = list()
        for key in data:
            row.append(data[key])

        for col_num in range(0,len(row)):
            c = ws.cell(row=row_num + 4, column=col_num + 2)
            c.value = row[col_num]
            c.alignment = Alignment(horizontal ='center',vertical ='center')

    wb.save(response)
    return response
